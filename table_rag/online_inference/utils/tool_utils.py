"""
tool_utils.py

POC adaptation: the original Embedder/Reranker used local bge-m3 / bge-reranker-v2-m3
via torch+transformers on CUDA. That is exactly the part that is impossible with an
OpenAI-compatible *gateway* (no local model, no GPU), so it is the only thing swapped:
  - Embedder.encode  -> gateway embeddings (poc_eval.common.llm_gateway.embed)
  - Reranker         -> pass-through (no gateway rerank endpoint); preserves recall order
Everything else in the repo (retriever flow, agent loop, prompts, NL2SQL) is unchanged.
Heavy imports are made lazy so the module loads without torch/transformers/openpyxl.
"""

import sys
from typing import Union, List, Tuple
import numpy as np


def sigmoid(x):
    return 1 / (1 + np.exp(-x))


class Embedder:
    """Gateway-backed drop-in for the original bge-m3 embedder."""

    def __init__(self, model_path=None, device_id=1) -> None:
        self.model_path = model_path  # kept for signature compatibility; unused

    def encode(self, texts):
        from poc_eval.common import llm_gateway
        if isinstance(texts, str):
            texts = [texts]
        vecs = llm_gateway.embed(list(texts))
        return np.asarray(vecs, dtype="float32")


class Reranker:
    """Pass-through reranker (no gateway rerank endpoint).

    compute_score returns descending scores in input order, so the caller's
    `sorted(zip(scores, docs, ...), reverse=True)` preserves the recall ranking.
    """

    def __init__(self, model_name_or_path: str = None, *args, **kwargs) -> None:
        self.model_name_or_path = model_name_or_path

    def compute_score(self, sentence_pairs, batch_size: int = 256,
                      max_length: int = 512, normalize: bool = False) -> List[float]:
        if sentence_pairs and isinstance(sentence_pairs[0], str):
            sentence_pairs = [sentence_pairs]
        n = len(sentence_pairs)
        return [float(n - i) for i in range(n)]


def excel_to_markdown(file_path):
    from openpyxl import load_workbook  # lazy: only needed if reading real .xlsx
    workbook = load_workbook(file_path)

    content = ""
    file_name = file_path.split("/")[-1]
    table_name = file_name.replace(".xlsx", "")
    content += f"Table name: {table_name}\n"
    for sheet_name in workbook.sheetnames:
        work_sheet = workbook[sheet_name]
        for i, row in enumerate(work_sheet):
            columns = [str(c.value) for c in row if c.value is not None]
            content += " | " + " | ".join(columns) + " | \n"
            if i == 0:
                content += " | " + " | ".join(["---"] * len(columns)) + " | \n"
    return content


if __name__ == '__main__':
    print(excel_to_markdown("./test.xlsx"))
