import sys
from openpyxl import load_workbook
import torch
from typing import Union, List, Tuple
from transformers import AutoTokenizer, AutoModel, AutoModelForSequenceClassification
from tqdm import tqdm
import numpy as np

def sigmoid(x) :
    return 1 / (1 + np.exp(-x))

class Embedder :
    def __init__(self, model_path, device_id=1) -> None:
        self.model_path = model_path
        self.device = torch.device(f"cuda:{device_id}")
        self.tokenizer = AutoTokenizer.from_pretrained(self.model_path, use_fast=True)
        self.model = AutoModel.from_pretrained(self.model_path)
        self.model = self.model.to(self.device)
        self.model.eval()
    
    @torch.no_grad()
    def encode(self, texts) :
        features = self.tokenizer(texts, padding=True, truncation=True, 
                                    return_tensors="pt").to(self.device)
        model_output = self.model(**features)
        embs = model_output[0][:, 0].cpu().numpy()
        return embs

class Reranker :
    def __init__(
        self,
        model_name_or_path: str = None,
        use_fp16: bool = False,
        inference_mode: str = "huggingface",
        cache_dir: str = None,
        device: Union[str, int] = 4
    ) -> None:

        self.interence_mode = inference_mode
        self.tokenizer = AutoTokenizer.from_pretrained(model_name_or_path, cache_dir=cache_dir)

        if device and isinstance(device, str) :
            self.device = torch.device(device)
            if device == "auto" :
                use_fp16 = False
        else :
            if torch.cuda.is_available() :
                if device is not None :
                    self.device = torch.device(f"cuda:{device}")
                else :
                    self.device = torch.device("cuda")       

        self.model = AutoModelForSequenceClassification.from_pretrained(
            model_name_or_path,
            cache_dir=cache_dir,
            trust_remote_code=True
        )

        if use_fp16 :
            self.model.half()
        self.model.eval()

        self.model = self.model.to(self.device)

        if device is None :
            self.num_gpus = torch.cuda.device_count()
            if self.num_gpus > 1 :
                print(f"----- using {self.num_gpus}*GPUs -------")
                self.model = torch.nn.DataParallel(self.model)
        else :
            self.num_gpus = 1

    @torch.no_grad()
    def compute_score(self, sentence_paris: Union[List[Tuple[str, str]], Tuple[str, str]], batch_size: int = 256,
                        max_length: int = 512, normalize: bool = False) -> List[float] :
        if self.num_gpus > 0 :
            batch_size = batch_size * self.num_gpus
        
        assert isinstance(sentence_paris, list)
        if isinstance(sentence_paris[0], str) :
            sentence_paris = [sentence_paris]
        
        all_scores = []
        flag = False
        error_count = 0
        while not flag :
            try :
                test_inputs_batch = self.tokenizer(
                    sentence_paris[: min(len(sentence_paris), batch_size)],
                    padding=True,
                    truncation=True,
                    max_length=max_length,
                    return_tensors="pt"
                ).to(self.device)

                scores = self.model(**test_inputs_batch, return_dict=True).logits.view(-1).float()
                all_scores.extend(scores.cpu().numpy().tolist())
                flag = True

            except RuntimeError as e :
                batch_size = batch_size // 2
                error_count += 1
                print("adjust", batch_size)
            
            except torch.cuda.OutOfMemoryError as e :
                batch_size = batch_size // 2
                error_count += 1
                print("adjust", batch_size)
            
            finally :
                if error_count > 5 :
                    raise NotImplementedError('error count')
            
        for start_index in tqdm(range(batch_size, len(sentence_paris), batch_size), desc="Compute Scores",
                                disable=len(sentence_paris) < batch_size) :
            sentences_batch = sentence_paris[start_index: start_index + batch_size]

            inputs = self.tokenizer(
                sentences_batch,
                padding=True,
                truncation=True,
                return_tensors="pt",
                max_length=max_length
            ).to(self.device)

            scores = self.model(**inputs, return_dict=True).logits.view(-1, ).float()
            all_scores.extend(scores.cpu().numpy().tolist())

        if normalize :
            all_scores = [sigmoid(score) for score in all_scores]
        
        return all_scores
    
def excel_to_markdown(file_path) :
    workbook = load_workbook(file_path)

    content = ""
    file_name = file_path.split("/")[-1]
    table_name = file_name.replace(".xlsx", "")
    content += f"Table name: {table_name}\n"
    for sheet_name in workbook.sheetnames :
        work_sheet = workbook[sheet_name]
        row_count = 0
        for i, row in enumerate(work_sheet) :
            columns = []
            for column in row :
                if column.value is None :
                    continue
                columns.append(column.value)

            content += " | " + " | ".join(columns) + " | \n"
            if i == 0 :
                content += " | " + " | ".join(["---"]*len(columns)) + " | \n"
            row_count += 1
    return content


if __name__ == '__main__' :
    print(excel_to_markdown("./test.xlsx"))