"""
Table -> Markdown rendering.

Re-implements the repo's `tool_utils.excel_to_markdown` (the repo version sits in a
module that imports torch/transformers at import time, which we deliberately avoid --
all model inference goes through the gateway instead).
"""
from __future__ import annotations

import os

from openpyxl import load_workbook


def excel_to_markdown(file_path: str) -> str:
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    table_name = os.path.basename(file_path).replace(".xlsx", "")
    content = f"Table name: {table_name}\n"
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            cells = [str(c) for c in row if c is not None]
            if not cells:
                continue
            content += " | " + " | ".join(cells) + " | \n"
            if i == 0:
                content += " | " + " | ".join(["---"] * len(cells)) + " | \n"
    workbook.close()
    return content
